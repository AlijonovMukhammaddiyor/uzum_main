import datetime
import io
import time
import traceback

import pandas as pd
import pytz
import requests
from django.db.models import OuterRef, Subquery
from django.utils import timezone
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from uzum.product.models import ProductAnalytics, ProductAnalyticsView
from uzum.shop.models import Shop, ShopAnalytics
from uzum.users.models import User
from uzum.utils.general import get_today_pretty, get_today_pretty_fake


def prepare_shop_statistics(user, shop: Shop):
    try:
        print(f"Shop Daily Sales Report for shop: {shop.title} and user: {user.username}")
        today_pretty = get_today_pretty()
        date = today_pretty

        start_date = timezone.make_aware(
            datetime.datetime.strptime(date, "%Y-%m-%d"), timezone=pytz.timezone("Asia/Tashkent")
        ).replace(hour=0, minute=0, second=0, microsecond=0)

        def calculate_diff(target, before):
            """
            Helper function to calculate the difference between the target and before value.
            """
            if target is not None and before is not None:
                return target - before
            return target

        # analytics for today
        analytics_date = (
            ProductAnalyticsView.objects.filter(shop_link=shop.link)
            .values(
                "product_id",
                "product_title",
                "product_title_ru",
                "category_title",
                "category_title_ru",
                "orders_amount",
                "diff_orders_amount",
                "weekly_orders_amount",
                "orders_money",
                "diff_orders_money",
                "weekly_orders_money",
                "reviews_amount",
                "diff_reviews_amount",
                "weekly_reviews_amount",
                "avg_purchase_price",
                "rating",
                "position_in_category",
                "product_available_amount",
            )
            .order_by("-orders_amount")
        )

        latest_date_subquery = (
            ProductAnalytics.objects.filter(
                product__shop=shop, created_at__lt=start_date, product_id=OuterRef("product_id")
            )
            .order_by("-created_at")
            .values("date_pretty")[:1]
        )

        day_before_analytics = ProductAnalytics.objects.filter(
            product__shop=shop, date_pretty=Subquery(latest_date_subquery)
        ).values(
            "average_purchase_price",
            "orders_amount",
            "product__product_id",
            "position_in_category",
            "available_amount",
            "reviews_amount",
            "rating",
        )

        target_analytics = list(analytics_date)

        before_analytics_dict = {i["product__product_id"]: i for i in day_before_analytics}

        for item in target_analytics:
            before_item = before_analytics_dict.get(item["product_id"], None)

            item["orders_count_yesterday"] = (
                calculate_diff(item["orders_amount"], before_item["orders_amount"])
                if before_item
                else item["orders_amount"]
            )

            item["reviews_count_yesterday"] = (
                calculate_diff(item["reviews_amount"], before_item["reviews_amount"])
                if before_item
                else item["reviews_amount"]
            )

            item["rating_yesterday"] = (
                calculate_diff(item["rating"], before_item["rating"]) if before_item else item["rating"]
            )

            item["position_in_category_yesterday"] = (
                calculate_diff(item["position_in_category"], before_item["position_in_category"])
                if before_item
                else None
            )

            item["price_yesterday"] = (
                calculate_diff(item["avg_purchase_price"], before_item["average_purchase_price"])
                if before_item
                else None
            )

            item["available_amount_yesterday"] = (
                calculate_diff(item["product_available_amount"], before_item["available_amount"])
                if before_item
                else item["product_available_amount"]
            )

        desired_order = [
            "product_id",
            "product_title",
            "product_title_ru",
            "category_title",
            "category_title_ru",
            "weekly_orders_amount",
            "orders_count_yesterday",
            "weekly_orders_money",
            "reviews_amount",
            "weekly_reviews_amount",
            "reviews_count_yesterday",
            "avg_purchase_price",
            "rating",
        ]
        target_analytics = [{key: item[key] for key in desired_order} for item in target_analytics]

        # for each of orders_money, diff_orders_money, weekly_orders_money, multiply by 1000
        for item in target_analytics:
            for key in ["orders_money", "diff_orders_money", "weekly_orders_money"]:
                if key in item and item[key] is not None:
                    item[key] = round(item[key]) * 1000

        for item in target_analytics:
            for key in ["avg_purchase_price", "price_yesterday"]:
                if key in item and item[key] is not None:
                    item[key] = (round(item[key]) / 100) * 100

        final_res = target_analytics

        return final_res
    except Exception as e:
        print("Error in ShopDailySalesView: ", e)
        traceback.print_exc()
        return None


def apply_color_based_on_value(cell, min_val, max_val):
    """Apply background color based on its value."""
    try:
        value = float(cell.value)

        if value == 0:
            return

        if value >= 0:
            if max_val == 0:  # Avoid division by zero
                ratio = 0
            else:
                # Use a power function to make small differences more pronounced
                ratio = (value / max_val) ** 2.5

            # Ensure ratio is between 0 and 1
            ratio = max(0, min(1, ratio))

            # Interpolating between 255 (FF) and 170 (AA) for green
            green_value = int(255 - (85 * ratio))
            fill = PatternFill(
                start_color=f"00{green_value:02X}00", end_color=f"00{green_value:02X}00", fill_type="solid"
            )
            cell.fill = fill
        else:
            if abs(min_val) == 0:  # Avoid division by zero
                ratio = 0
            else:
                # For negative values, we'll calculate the ratio differently
                ratio = (abs(value) / abs(min_val)) ** 2.5

            # Ensure ratio is between 0 and 1
            ratio = max(0, min(1, ratio))

            # Interpolating between 255 (FF) and 210 (D2) for red
            red_value = int(255 - (45 * ratio))
            # Adding a bit of blue and green to make it pinkish, but ensuring it doesn't exceed 255
            green_blue_value = int(210 * (1 - ratio))
            fill = PatternFill(
                start_color=f"{red_value:02X}{green_blue_value:02X}{green_blue_value:02X}",
                end_color=f"{red_value:02X}{green_blue_value:02X}{green_blue_value:02X}",
                fill_type="solid",
            )
            cell.fill = fill
    except Exception as e:
        pass


def export_to_excel(shops_data, products_data, user):
    try:
        # Define the path to save the Excel file
        output = io.BytesIO()

        # Column name mapping
        column_mapping = {
            "product_id": "ID ПРОДУКТА",
            "product_title": "НАЗВАНИЕ ПРОДУКТА",
            "product_title_ru": "НАЗВАНИЕ ПРОДУКТА RU",
            "category_title": "НАЗВАНИЕ КАТЕГОРИИ",
            "category_title_ru": "НАЗВАНИЕ КАТЕГОРИИ RU",
            "orders_amount": "КОЛИЧЕСТВО ЗАКАЗОВ",
            "product_available_amount": "НАЛИЧИЕ ТОВАРА",
            "reviews_amount": "КОЛИЧЕСТВО ОТЗЫВОВ",
            "rating": "РЕЙТИНГ",
            "position_in_category": "ПОЗИЦИЯ В КАТЕГОРИИ",
            "avg_purchase_price": "СРЕДНЯЯ ЦЕНА",
            # orders_money: total revenue общая выручка
            "orders_money": "ОБЩАЯ ВЫРУЧКА",
            "diff_orders_amount": "КОЛИЧЕСТВЕ ЗАКАЗОВ (30 ДНЕЙ)",
            "diff_reviews_amount": "КОЛИЧЕСТВЕ ОТЗЫВОВ (30 ДНЕЙ)",
            "diff_orders_money": "ОБЩАЯ ВЫРУЧКА (30 ДНЕЙ)",
            "weekly_orders_money": "ЕЖЕНЕДЕЛЬНАЯ ВЫРУЧКА",
            "weekly_orders_amount": "ЕЖЕНЕДЕЛЬНОЕ КОЛИЧЕСТВО ЗАКАЗОВ",
            "weekly_reviews_amount": "ЕЖЕНЕДЕЛЬНОЕ КОЛИЧЕСТВО ОТЗЫВОВ",
            "orders_count_yesterday": "КОЛИЧЕСТВО ЗАКАЗОВ ВЧЕРА",
            "reviews_count_yesterday": "КОЛИЧЕСТВО ОТЗЫВОВ ВЧЕРА",
            "rating_yesterday": "ИЗМЕНЕНИЕ РЕЙТИНГА ВЧЕРА",
            "position_in_category_yesterday": "ИЗМЕНЕНИЕ ПОЗИЦИИ В КАТЕГОРИИ ВЧЕРА",
            "price_yesterday": "ИЗМЕНЕНИЕ ЦЕНЫ ВЧЕРА",
            "available_amount_yesterday": "ИЗМЕНЕНИЕ ДОСТУПНОГО КОЛИЧЕСТВА ВЧЕРА",
        }

        shop_column_mapping = {
            "total_products": "КОЛИЧЕСТВО ПРОДУКТОВ",
            "total_orders": "КОЛИЧЕСТВО ЗАКАЗОВ",
            "total_revenue": "ОБЩАЯ ВЫРУЧКА",
            "total_reviews": "КОЛИЧЕСТВО ОТЗЫВОВ",
            "average_purchase_price": "СРЕДНЯЯ ЦЕНА ПРОДУКТА",
            "rating": "РЕЙТИНГ",
            # Позиция на основе дохода
            "position": "ПОЗИЦИЯ",
            "categories": "КОЛИЧЕСТВО КАТЕГОРИЙ",
        }

        product_column_mapping = {
            "product_id": "ID ПРОДУКТА",
            "product_title": "НАЗВАНИЕ ПРОДУКТА",
            "product_title_ru": "НАЗВАНИЕ ПРОДУКТА RU",
            "category_title": "НАЗВАНИЕ КАТЕГОРИИ",
            "category_title_ru": "НАЗВАНИЕ КАТЕГОРИИ RU",
            "orders_amount": "КОЛИЧЕСТВО ЗАКАЗОВ",
            "diff_orders_amount": "КОЛИЧЕСТВЕ ЗАКАЗОВ (30 ДНЕЙ)",
            "weekly_orders_amount": "ЕЖЕНЕДЕЛЬНОЕ КОЛИЧЕСТВО ЗАКАЗОВ",
            "orders_money": "ОБЩАЯ ВЫРУЧКА",
            "diff_orders_money": "ОБЩАЯ ВЫРУЧКА (30 ДНЕЙ)",
            "weekly_orders_money": "ЕЖЕНЕДЕЛЬНАЯ ВЫРУЧКА",
            "reviews_amount": "КОЛИЧЕСТВО ОТЗЫВОВ",
            "diff_reviews_amount": "КОЛИЧЕСТВЕ ОТЗЫВОВ (30 ДНЕЙ)",
            "weekly_reviews_amount": "ЕЖЕНЕДЕЛЬНОЕ КОЛИЧЕСТВО ОТЗЫВОВ",
            "avg_purchase_price": "СРЕДНЯЯ ЦЕНА",
            "rating": "РЕЙТИНГ",
            "position_in_category": "ПОЗИЦИЯ В КАТЕГОРИИ",
            "product_available_amount": "НАЛИЧИЕ ТОВАРА",
        }

        # Create a new Excel writer object
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for shop_name, data in shops_data.items():
                # Convert shop analytics to a pandas DataFrame
                shop_df = pd.DataFrame([data["shop_analytics"]])

                shop_df.rename(columns=shop_column_mapping, inplace=True)

                # Convert products analytics to a pandas DataFrame
                products_df = pd.DataFrame(data["products_analytics_of_shop"])

                # Rename columns
                products_df.rename(columns=column_mapping, inplace=True)

                # Write data to Excel
                valid_sheet_name = (
                    shop_name[:31]
                    .replace(":", "")
                    .replace("\\", "")
                    .replace("/", "")
                    .replace("?", "")
                    .replace("*", "")
                    .replace("[", "")
                    .replace("]", "")
                )
                shop_df.to_excel(writer, sheet_name=valid_sheet_name, startrow=0, index=False)

                products_df.to_excel(writer, sheet_name=valid_sheet_name, startrow=len(shop_df) + 2, index=False)

                # Apply styling using openpyxl
                ws = writer.sheets[valid_sheet_name]
                for row in ws.iter_rows():
                    ws.row_dimensions[row[0].row].height = 20

                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            # Check the length of the cell value and the header (cell in the first row)
                            max_length = max(max_length, len(str(cell.value)), len(str(column[0].value)))
                            cell.font = Font(size=12)
                        except Exception as e:
                            pass
                    adjusted_width = max_length + 2  # Add 2 for a little extra space
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                # Apply conditional formatting for columns with numbers
                for col in ws.columns:
                    col_values = [
                        cell.value for cell in col if cell.value is not None and isinstance(cell.value, (int, float))
                    ]
                    if col_values:
                        min_val, max_val = min(col_values), max(col_values)
                        for cell in col:
                            if cell.column_letter != "A":  # Assuming product_id is in the first column
                                apply_color_based_on_value(cell, min_val, max_val)

                # Apply borders to all cells
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border

            # all products
            ws = writer.book.create_sheet(title="ВСЕ ПРОДУКТЫ")

            # Define colors for headers, data, and alternating product cards
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )  # Dark blue for headers
            data_fill_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
            data_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

            header_font = Font(size=14, bold=True, color="FFFFFF")  # White font for headers
            data_font = Font(size=12)
            positive_fill = PatternFill(
                start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"
            )  # Greenish for positive changes
            negative_fill = PatternFill(
                start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
            )  # Pinkish for negative changes

            row_num = 1
            color_toggle = True
            row_num = 1
            for product_id, data in products_data.items():
                current_data = data["current"]

                if not current_data:
                    continue

                latest_analytics = data["latest_analytics"]

                current_fill = data_fill_1 if color_toggle else data_fill_2
                color_toggle = not color_toggle

                # id
                ws[f"A{row_num}"] = "ID ПРОДУКТА"
                ws[f"A{row_num}"].fill = header_fill
                ws[f"A{row_num}"].font = header_font
                ws[f"A{row_num}"].alignment = Alignment(horizontal="center")
                ws[f"B{row_num}"] = current_data["product_id"]
                ws[f"B{row_num}"].fill = current_fill
                ws[f"B{row_num}"].font = data_font
                row_num += 1

                # Product Title
                ws[f"A{row_num}"] = "НАЗВАНИЕ ПРОДУКТА UZ"
                ws[f"B{row_num}"] = current_data["product_title"]
                row_num += 1

                # Product Title (RU)
                ws[f"A{row_num}"] = "НАЗВАНИЕ ПРОДУКТА RU"
                ws[f"B{row_num}"] = current_data["product_title_ru"] if current_data["product_title_ru"] else "-"
                row_num += 1

                # category
                ws[f"A{row_num}"] = "НАЗВАНИЕ КАТЕГОРИИ UZ"
                ws[f"B{row_num}"] = current_data["category_title"]
                row_num += 1

                # category (RU)
                ws[f"A{row_num}"] = "НАЗВАНИЕ КАТЕГОРИИ RU"
                ws[f"B{row_num}"] = current_data["category_title_ru"] if current_data["category_title_ru"] else "-"
                row_num += 1

                # shop
                ws[f"A{row_num}"] = "НАЗВАНИЕ МАГАЗИНА"
                ws[f"B{row_num}"] = current_data["shop_title"]
                row_num += 1

                # # Orders
                # ws[f"A{row_num}"] = "КОЛИЧЕСТВО ЗАКАЗОВ"
                # ws[f"B{row_num}"] = current_data["orders_amount"]
                # row_num += 1

                ws[f"A{row_num}"] = "КОЛИЧЕСТВЕ ЗАКАЗОВ (30 ДНЕЙ)"
                ws[f"B{row_num}"] = current_data["diff_orders_amount"]

                row_num += 1

                ws[f"A{row_num}"] = "ЕЖЕНЕДЕЛЬНОЕ КОЛИЧЕСТВО ЗАКАЗОВ (7 ДНЕЙ)"
                ws[f"B{row_num}"] = current_data["weekly_orders_amount"]
                row_num += 1

                # number od orders yesterday
                ws[f"A{row_num}"] = "КОЛИЧЕСТВО ЗАКАЗОВ ВЧЕРА"
                ws[f"B{row_num}"] = current_data["orders_amount"] - latest_analytics["orders_amount"]
                row_num += 1

                # # orders_money: total revenue общая выручка
                # ws[f"A{row_num}"] = "ОБЩАЯ ВЫРУЧКА"
                # ws[f"B{row_num}"] = round(current_data["orders_money"]) * 1000
                # row_num += 1

                ws[f"A{row_num}"] = "ОБЩАЯ ВЫРУЧКА (30 ДНЕЙ)"
                ws[f"B{row_num}"] = round(current_data["diff_orders_money"]) * 1000
                row_num += 1

                ws[f"A{row_num}"] = "ЕЖЕНЕДЕЛЬНАЯ ВЫРУЧКА (7 ДНЕЙ)"
                ws[f"B{row_num}"] = round(current_data["weekly_orders_money"]) * 1000
                row_num += 1

                # # Reviews
                # ws[f"A{row_num}"] = "КОЛИЧЕСТВО ОТЗЫВОВ"
                # ws[f"B{row_num}"] = current_data["reviews_amount"]
                # row_num += 1

                ws[f"A{row_num}"] = "КОЛИЧЕСТВЕ ОТЗЫВОВ (30 ДНЕЙ)"
                ws[f"B{row_num}"] = current_data["diff_reviews_amount"]
                row_num += 1

                ws[f"A{row_num}"] = "ЕЖЕНЕДЕЛЬНОЕ КОЛИЧЕСТВО ОТЗЫВОВ (7 ДНЕЙ)"
                ws[f"B{row_num}"] = current_data["weekly_reviews_amount"]
                row_num += 1

                ws[f"A{row_num}"] = "КОЛИЧЕСТВО ОТЗЫВОВ ВЧЕРА"
                ws[f"B{row_num}"] = current_data["reviews_amount"] - latest_analytics["reviews_amount"]
                row_num += 1

                # Rating
                ws[f"A{row_num}"] = "РЕЙТИНГ:"
                ws[f"B{row_num}"] = current_data["rating"]
                row_num += 1

                # ws[f"A{row_num}"] = "ИЗМЕНЕНИЕ РЕЙТИНГА ВЧЕРА:"
                # ws[f"B{row_num}"] = current_data["rating"] - latest_analytics["rating"]
                # change_rating_cell = ws[f"B{row_num}"]
                # if change_rating_cell.value < 0:
                #     change_rating_cell.fill = negative_fill
                # elif change_rating_cell.value > 0:
                #     change_rating_cell.fill = positive_fill
                # row_num += 1

                # # Position in Category
                # ws[f"A{row_num}"] = "ПОЗИЦИЯ В КАТЕГОРИИ:"
                # ws[f"B{row_num}"] = current_data["position_in_category"]
                # row_num += 1

                # ws[f"A{row_num}"] = "ИЗМЕНЕНИЕ ПОЗИЦИИ В КАТЕГОРИИ ВЧЕРА:"
                # ws[f"B{row_num}"] = current_data["position_in_category"] - latest_analytics["position_in_category"]
                # change_position_cell = ws[f"B{row_num}"]
                # if change_position_cell.value < 0:
                #     change_position_cell.fill = negative_fill
                # elif change_position_cell.value > 0:
                #     change_position_cell.fill = positive_fill
                # row_num += 1

                # Average Purchase Price
                ws[f"A{row_num}"] = "СРЕДНЯЯ ЦЕНА:"
                ws[f"B{row_num}"] = round(round(current_data["avg_purchase_price"]) / 100) * 100
                row_num += 1

                ws[f"A{row_num}"] = "ИЗМЕНЕНИЕ ЦЕНЫ ВЧЕРА:"
                ws[f"B{row_num}"] = round(round(current_data["avg_purchase_price"]) / 100) * 100 - (
                    round(round(latest_analytics["average_purchase_price"]) / 100) * 100
                )
                change_price_cell = ws[f"B{row_num}"]

                if change_price_cell.value < 0:
                    change_price_cell.fill = negative_fill
                elif change_price_cell.value > 0:
                    change_price_cell.fill = positive_fill

                row_num += 1

                # Available Amount
                ws[f"A{row_num}"] = "НАЛИЧИЕ ТОВАРА:"
                ws[f"B{row_num}"] = current_data["product_available_amount"]

                row_num += 4

            for column in ws.columns:
                for cell in column:
                    cell.border = Border(
                        left=Side(style="double"),
                        right=Side(style="double"),
                        top=Side(style="double"),
                        bottom=Side(style="double"),
                    )

            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        output.seek(0)

        return output
    except Exception as e:
        print("Error in export_to_excel: ", e)
        traceback.print_exc()
        return None


def send_reports_to_all():
    try:
        users = User.objects.filter(is_telegram_connected=True).exclude(tariff="free")

        for user in users:
            # Generate the report for the user
            favourite_shops = user.favourite_shops.all()
            favourite_products = user.favourite_products.all()

            all_shop_data = {}

            for shop in favourite_shops:
                # Fetch the latest analytics for the shop
                latest_analytics = ShopAnalytics.objects.filter(shop=shop).order_by("-created_at").first()

                # Convert the ShopAnalytics object to a dictionary
                shop_analytics_dict = {
                    "total_products": latest_analytics.total_products,
                    "total_orders": latest_analytics.total_orders,
                    "total_revenue": round(latest_analytics.total_revenue) * 1000,
                    "total_reviews": latest_analytics.total_reviews,
                    "average_purchase_price": latest_analytics.average_purchase_price,
                    "rating": latest_analytics.rating,
                    "categories": latest_analytics.categories.all().count(),
                }

                # Fetch the daily product analytics for the shop
                shop_data = prepare_shop_statistics(user, shop)

                all_shop_data[shop.title] = {
                    "shop_analytics": shop_analytics_dict,
                    "products_analytics_of_shop": shop_data,
                }

            all_products_data = {}

            all_products_current = (
                ProductAnalyticsView.objects.filter(
                    product_id__in=favourite_products.values_list("product_id", flat=True)
                )
                .values(
                    "product_id",
                    "product_title",
                    "product_title_ru",
                    "category_title",
                    "category_title_ru",
                    "shop_title",
                    "shop_title",
                    "orders_amount",
                    "diff_orders_amount",
                    "weekly_orders_amount",
                    "orders_money",
                    "diff_orders_money",
                    "weekly_orders_money",
                    "reviews_amount",
                    "diff_reviews_amount",
                    "weekly_reviews_amount",
                    "avg_purchase_price",
                    "rating",
                    "position_in_category",
                    "product_available_amount",
                )
                .order_by("-orders_amount")
            )

            for product in favourite_products:
                analytics = (
                    ProductAnalytics.objects.filter(product=product)
                    .order_by("-created_at")
                    .values(
                        "orders_amount",
                        "reviews_amount",
                        "position_in_category",
                        "orders_money",
                        "average_purchase_price",
                        "available_amount",
                        "date_pretty",
                    )
                )

                if analytics.count() > 1:
                    # get the analytics for the previous day
                    latest_analytics = analytics[1]
                    # change the orders_money by multiplying by 1000
                if product:
                    all_products_data[product.product_id] = {
                        "date_pretty": get_today_pretty(),
                        "current": all_products_current.filter(product_id=product.product_id).first(),
                        "latest_analytics": latest_analytics,
                    }

            # Export the combined shop data to Excel
            file_path = export_to_excel(all_shop_data, all_products_data, user)

            send_file_to_telegram_bot(user.telegram_chat_id, file_path)

    except Exception as e:
        print("Error in UserDailyReport: ", e)
        return None


def send_file_to_telegram_bot(chat_id, file_stream):
    bot_token = "6419033506:AAETG8prNWtydbqFEdiiFa-z_YxRaRSbzA8"
    send_document_url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    date_pretty = get_today_pretty()
    data = {"chat_id": int(chat_id), "caption": "Вот ваш ежедневный отчет. Этот файл Excel содержит несколько листов."}
    files = {
        "document": (
            f"отчет-{date_pretty}.xlsx",
            file_stream,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }
    response = requests.post(send_document_url, data=data, files=files)
    return response.json()


def send_to_single_user(user):
    try:
        # Generate the report for the user
        favourite_shops = user.favourite_shops.all()
        favourite_products = user.favourite_products.all()

        all_shop_data = {}

        for shop in favourite_shops:
            # Fetch the latest analytics for the shop
            latest_analytics = ShopAnalytics.objects.filter(shop=shop).order_by("-created_at").first()

            # Convert the ShopAnalytics object to a dictionary
            shop_analytics_dict = {
                "total_products": latest_analytics.total_products,
                "total_orders": latest_analytics.total_orders,
                "total_revenue": round(latest_analytics.total_revenue) * 1000,
                "total_reviews": latest_analytics.total_reviews,
                "average_purchase_price": latest_analytics.average_purchase_price,
                "rating": latest_analytics.rating,
                "categories": latest_analytics.categories.all().count(),
            }

            # Fetch the daily product analytics for the shop
            shop_data = prepare_shop_statistics(user, shop)

            all_shop_data[shop.title] = {
                "shop_analytics": shop_analytics_dict,
                "products_analytics_of_shop": shop_data,
            }

        all_products_data = {}

        all_products_current = (
            ProductAnalyticsView.objects.filter(product_id__in=favourite_products.values_list("product_id", flat=True))
            .values(
                "product_id",
                "product_title",
                "product_title_ru",
                "category_title",
                "category_title_ru",
                "shop_title",
                "orders_amount",
                "diff_orders_amount",
                "weekly_orders_amount",
                "orders_money",
                "diff_orders_money",
                "weekly_orders_money",
                "reviews_amount",
                "diff_reviews_amount",
                "weekly_reviews_amount",
                "avg_purchase_price",
                "rating",
                "position_in_category",
                "product_available_amount",
            )
            .order_by("-orders_amount")
        )

        for product in favourite_products:
            analytics = (
                ProductAnalytics.objects.filter(product=product)
                .order_by("-created_at")
                .values(
                    "orders_amount",
                    "reviews_amount",
                    "position_in_category",
                    "orders_money",
                    "rating",
                    "average_purchase_price",
                    "available_amount",
                    "date_pretty",
                )
            )

            if analytics.count() > 1:
                # get the analytics for the previous day
                latest_analytics = analytics[1]

            if product:
                all_products_data[product.product_id] = {
                    "date_pretty": get_today_pretty(),
                    "current": all_products_current.filter(product_id=product.product_id).first(),
                    "latest_analytics": latest_analytics,
                }

        # Export the combined shop data to Excel
        file_path = export_to_excel(all_shop_data, all_products_data, user)

        send_file_to_telegram_bot(user.telegram_chat_id, file_path)

    except Exception as e:
        print("Error in Sending to single user: ", e)
        return None
